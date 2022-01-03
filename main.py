import datetime
import sys
import time
import traceback
import os
from MyWindows import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QDate,QEvent
from PyQt5.QtGui import QFont,QIcon
# import xlwt
# import xlrd
import openpyxl
# from xlutils.copy import copy
from xlsxwriter import Workbook

add_count = 0
ouput_excel = "patient_info_temp.xlsx"
isExistPatient = False
inserRow = 1
updateRow = 1
patient_title = ['病人ID号','姓名','性别','年龄','出生日期','检查日期','诊断','其他','息肉个数','息肉大小','息肉部位','内镜表现','息肉病理诊断','癌灶个数','病理：按分化程度','病理：按形态分类','内镜形态','病理：按组织来源','部位']
patient_key = ["ID","Name","Sex","Age","BirthDay","CheckDate","Diagnose","OtherDia","PolyoCount",
               "PolyoSize","PolyoSite","Endoscope","PolyoPathology","CancerFociCount","DifferePathology",
               "ShapePathology","EndoscopeShape","HistologicPathology","CancerSite"]
patient_info = {"ID":0,
                "Name":"",
                "Sex":"男",
                "Age":0,
                "BirthDay":"",
                "CheckDate":"",
                "Diagnose":"",
                "OtherDia":"",
                "PolyoCount":"",
                "PolyoSize":"",
                "PolyoSite":"",
                "Endoscope":"",
                "PolyoPathology":"",
                "CancerFociCount":"",
                "DifferePathology":"",
                "ShapePathology":"",
                "EndoscopeShape":"",
                "HistologicPathology":"",
                "CancerSite":""
                }
# print(len(patient_info.keys()))
# print(patient_info.keys())


def writeExcel(path,dic_info,count=1):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    k_list = []
    v_list = []
    for key, value in dic_info.items():
        k_list.append(key)
        v_list.append(value)
        for i in range(0,len(k_list)):
            sheet.cell(row=1,column=i+1,value=k_list[i])
            for j in range(0,len(v_list)):
                sheet.cell(row=i+2,column=j+1,value=v_list[j])
    workbook.save(path)



class myWin(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super(myWin,self).__init__()
        self.setupUi(self)
        self.initWidgetsParm()
        self.tabItem = {"癌":self.tab,"息肉":self.Polyo,"正常":"","其他":self.tab_2}
        self.tabItemList = [{"正常": "","checkResult" : 0},{"息肉": self.Polyo,"checkResult" : 1},{"癌": self.tab,"checkResult" : 2},{"其他":"","checkResult" : 3}]
        self.isCheck = False
        self.addCount = 1
        self.read_patient_info = []
        # self.Sex = ["男","女"]
        # self.isCancer = False
        # self.isPolyo = False
        font = QFont('微软雅黑', 13)
        font.setBold(True)  # 设置字体加粗
        self.setWindowIcon(QIcon('./avc.ico'))
        self.tableWidget.horizontalHeader().setFont(font)
        self.tableWidget.setRowCount(add_count)
        self.tableWidget.setColumnCount(len(patient_title))
        self.tableWidget.setHorizontalHeaderLabels(patient_title)
        self.tableWidget.horizontalHeader().setStyleSheet("QHeaderView::section{border:2px groove gray;padding:2px 4px;background-color: #CDDEFF;}")
        # self.old_hook = sys.excepthook
        # sys.excepthook = self.catch_exceptions

        # self.retranslateUi(self)
        # PushSend = self.pushButton
        # qss = '''QPushButton{background-color:red;}'''
        # self.setStyleSheet(qss)
        #
        # self.setStyleSheet("QLineEdit#input { border:1px solid #0F0F0E;} QLineEdit#inputNull { border:1px solid #FF5959;}")
        self.pushButton.clicked.connect(self.sendPush)
        self.pushButton_2.clicked.connect(self.writePatienInfo)
        #根据出生年月设置来算年龄
        self.comboBox_6.currentTextChanged.connect(self.addTabItem)
        self.dateEdit.dateChanged.connect(self.calAge)
        self.tabWidget.tabCloseRequested.connect(self.closeTab)
        self.tabWidget.installEventFilter(self)
        # self.lineEdit_4.

    def catch_exceptions(self,ty,value,tb):
        trace_format = traceback.format_exception(ty,value,tb)
        print("ty:{}".format(ty))
        print("value:{}".format(value))
        print("tb:{}".format(tb))
        trace_string = "".join(trace_format)
        QMessageBox.critical(None,"An exception was raised","{}".format(trace_format))
        self.old_hook(ty,value,tb)

    def writePatienInfo(self):
        isExistPatient = False
        print("writePatienInfo")
        for index in range(len(self.read_patient_info)):
            if patient_info["Name"] == self.read_patient_info[index][0] and patient_info["BirthDay"] == self.read_patient_info[index][1]:
                print(index,self.read_patient_info[index])
                print("存在同一个人")
                isExistPatient = True
                updateRow = index
        infoList = []
        loadWb = openpyxl.load_workbook(ouput_excel)
        sheet1 = loadWb.get_sheet_by_name(loadWb.sheetnames[0])
        for k,v in patient_info.items():
            infoList.append(v)
        if not isExistPatient:
            sheet1.append(infoList)
        else:
            # sheet1.insert_rows(inserRow)
            # print(len(tuple(sheet1.rows)))
            # print(tuple(sheet1.rows))
            # print("updateRow:",updateRow)
            for index in range(len(sheet1[updateRow + 1])):
                # print(sheet1[updateRow + 1][index].value)
                sheet1[updateRow + 1][index].value = patient_info[patient_key[index]]
        loadWb.save(ouput_excel)

    def readInfoForSort(self):

        loadWb = openpyxl.load_workbook(ouput_excel,read_only=True)
        ws = loadWb[loadWb.sheetnames[0]]
        for row in ws.rows:
            row_info = []
            # print(row)
            index = 0
            for cell in row:
                if index > 4:
                    break
                if index == 1 or index == 4:
                    row_info.append(cell.value)
                index = index + 1
                # print("index:", index)
            if len(row_info):
                self.read_patient_info.append(row_info)

    def eventFilter(self, object, event):
        if event.type() == QEvent.Enter :
            print("mouse Enter ")
            self.setStyleSheet("QLineEdit#lineEdit_4,#lineEdit_3,QComboBox#comboBox_6 { border:1px solid #828790;}")
            return True
        return False

    def sendPush(self):
        print("SendPush has been on clicked")
        # print(self.comboBox.currentText())
        # print(self.comboBox_2.currentText())
        if self.lineEdit_4.text() == "":
            print("lineEdit_4空")
            self.setStyleSheet("QLineEdit#lineEdit_4 { border:1px solid #FF5959;background-color:#E4F0FA}")
            return

        if self.lineEdit_3.text() == "":
            print("lineEdit_3空")
            self.setStyleSheet("QLineEdit#lineEdit_3 { border:1px solid #FF5959;}")
            return

        if self.comboBox_6.currentText() == "":
            print("诊断信息未填")
            self.setStyleSheet("QComboBox#comboBox_6 { border:1px solid #FF5959;}")
            return
        elif self.comboBox_6.currentText() == "息肉":
            patient_info["PolyoSite"] = self.comboBox.currentText()
            patient_info["Endoscope"] = self.comboBox_2.currentText()
            patient_info["PolyoPathology"] = self.comboBox_3.currentText()
            patient_info["PolyoCount"] = self.comboBox_4.currentText()
            patient_info["PolyoSize"] = self.comboBox_7.currentText()
        elif self.comboBox_6.currentText() == "癌":
            patient_info["CancerFociCount"] = self.comboBox_12.currentText()
            patient_info["EndoscopeShape"] = self.comboBox_10.currentText()
            patient_info["DifferePathology"] = self.comboBox_13.currentText()
            patient_info["ShapePathology"] = self.comboBox_14.currentText()
            patient_info["HistologicPathology"] = self.comboBox_16.currentText()
            patient_info["CancerSite"] = self.comboBox_15.currentText()
        elif self.comboBox_6.currentText() == "其他":
            print("其他")
            patient_info["OtherDia"] = self.comboBox_5.currentText()

        patient_info["ID"] = self.lineEdit_4.text()
        patient_info["Name"] = self.lineEdit_3.text()
        patient_info["Age"] = self.comboBox_9.currentText()
        if self.femaleButton.isChecked():
            patient_info["Sex"] = "女性"
        elif self.maleButton.isChecked():
            patient_info["Sex"] = "男性"
        patient_info["BirthDay"] = self.dateEdit.date().toString("yyyy/MM/dd")
        patient_info["CheckDate"] = self.dateTimeEdit.dateTime().toString("yyyy/MM/dd hh:mm")
        patient_info["Diagnose"] = self.comboBox_6.currentText()

        print(patient_info)
        # self.addCount += 1


        try:
            self.read_patient_info = []
            self.readInfoForSort()
            print(len(self.read_patient_info))
            print(self.read_patient_info)
            self.writePatienInfo()

            rowcount = self.tableWidget.rowCount()
            self.tableWidget.setRowCount(rowcount + 1)
            for k, v in patient_info.items():
                # print(k,v)
                self.tableWidget.setItem(rowcount, patient_key.index(k), QTableWidgetItem(v))
        except PermissionError as err:
            print('Handling run-time error:', err)
            QMessageBox.critical(None, "文件打开错误", "{}被占用，请先关闭文件".format(ouput_excel))
            # self.tableWidget.removeRow(rowcount)

    def calAge(self):
        current_day = time.localtime(time.time())
        birth_day = self.dateEdit.date()
        print("计算年龄")
        age = 0
        print(str(self.dateEdit.date().year()) + "-" + str(self.dateEdit.date().month()) + "-" + str(self.dateEdit.date().day()))
        if self.dateEdit.date().year() < time.localtime(time.time()).tm_year :
            if current_day.tm_mon < birth_day.month() or (current_day.tm_mon == birth_day.month() and current_day.tm_mday < birth_day.day()):
                age = current_day.tm_year - birth_day.year() - 1
            if current_day.tm_mon > birth_day.month() or (current_day.tm_mon == birth_day.month() and current_day.tm_mday >= birth_day.day()):
                age = current_day.tm_year - birth_day.year()
        else:
            return
        self.comboBox_9.setCurrentText(str(age))

    def initWidgetsParm(self):
        self.dateEdit.setMaximumDate(QDate.currentDate())
        self.dateTimeEdit.setMaximumDate(QDate.currentDate())
        while(self.tabWidget.count() != 1):
            self.tabWidget.removeTab(1)
            self.tabWidget.removeTab(1)

    def addTabItem(self,tabTitle):
        print(tabTitle)
        if self.isCheck or tabTitle == "正常" or tabTitle == "":
            return
        self.tabWidget.addTab(self.tabItem[tabTitle], tabTitle)
        self.isCheck = True
        self.tabWidget.setCurrentIndex(1)

    def closeTab(self,item):
        print(item)
        print(self.tabWidget.tabText(self.tabWidget.currentIndex()))
        currentTabTitle = self.tabWidget.tabText(self.tabWidget.currentIndex())
        if currentTabTitle == "息肉":
            patient_info["PolyoSite"] = ""
            patient_info["Endoscope"] = ""
            patient_info["PolyoPathology"] = ""
            patient_info["PolyoCount"] = ""
            patient_info["PolyoSize"] = ""
        elif currentTabTitle == "癌":
            patient_info["CancerFociCount"] = ""
            patient_info["EndoscopeShape"] = ""
            patient_info["DifferePathology"] = ""
            patient_info["ShapePathology"] = ""
            patient_info["HistologicPathology"] = ""
            patient_info["CancerSite"] = ""
        elif currentTabTitle == "其他":
            patient_info["OtherDia"] = ""
        if self.tabWidget.count() == 1:
            return
        self.tabWidget.removeTab(item)
        self.isCheck = False
        self.comboBox_6.setCurrentIndex(0)

if __name__ == "__main__":
    if not os.path.exists(ouput_excel):
        workbook = Workbook(ouput_excel)  # 创建一个名为 hello.xlsx 赋值给workbook
        worksheet = workbook.add_worksheet("sheet1")
        for v in patient_title:
            worksheet.write(0,patient_title.index(v),v)
        workbook.close()
    app = QApplication(sys.argv)
    Win = myWin()
    Win.show()
    sys.exit(app.exec_())